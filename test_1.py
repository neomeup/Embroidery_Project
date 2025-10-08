import time

# Calculate the start time
start = time.time()


import openpyxl
from openpyxl.styles import PatternFill, Font
from PIL import Image
from colormath.color_objects import sRGBColor, LabColor
from colormath.color_conversions import convert_color
from colormath.color_diff import delta_e_cie2000
from sklearn.cluster import KMeans
import numpy as np

# User-defined parameters
INPUT_PATH = r"C:\Users\Neal\OneDrive\Desktop\pyproj\py3120\excel_sheet_brod\photo_3.jpg"
OUTPUT_PATH = "Julie_marion_example_50.xlsx"
GRID_ROWS = 50
GRID_COLS = 100
USE_ALL_COLORS = False  # Set to False to limit the palette
MAX_COLORS = 50         # Number of colors to use if USE_ALL_COLORS is False
FILTER_UNOWNED_COLORS = False  # Set to False to include unowned colors in the processed list

# Master list: [(RGB tuple), "Hex Code", "DMC code - Name", Owned (True/False)]
MASTER_LIST = [
    [(239, 238, 240), "F0EEEF", "1 - White Tin", False],
    [(197, 196, 201), "C9C4C5", "2 - Tin", False],
    [(176, 176, 181), "B5B0B0", "3 - Tin - Medium", False],
    [(156, 155, 157), "9D9B9C", "4 - Tin - Dark", False],
    [(227, 204, 190), "BECCE3", "5 - Driftwood - Light", False],
    [(220, 198, 184), "B8C6DC", "6 - Driftwood - Medium Light", False],
    [(204, 184, 170), "AAB8CC", "7 - Driftwood", False],
    [(157, 125, 113), "717D9D", "8 - Driftwood - Dark", False],
    [(85, 32, 20), "142055", "9 - Cocoa - Very Dark", False],
    [(237, 254, 217), "D9FEED", "10 - Tender Green - Very Light", False],
    [(226, 237, 181), "B5EDE2", "11 - Tender Green - Light", False],
    [(205, 217, 154), "9AD9CD", "12 - Tender Green", False],
    [(191, 246, 224), "E0F6BF", "13 - Nile Green - Medium Light", False],
    [(208, 251, 178), "B2FBD0", "14 - Apple Green - Pale", False],
    [(209, 237, 164), "A4EDD1", "15 - Apple Green", False],
    [(164, 214, 124), "7CD6A4", "16 - Chartreuse - Light", False],
    [(229, 226, 114), "72E2E5", "17 - Yellow Plum - Light", False],
    [(217, 213, 109), "6DD5D9", "18 - Yellow Plum", False],
    [(247, 201, 95), "5FC9F7", "19 - Autumn Gold - Medium Light", False],
    [(247, 175, 147), "93AFF7", "20 - Shrimp", False],
    [(215, 153, 130), "8299D7", "21 - Alizarian - Light", False],
    [(188, 96, 78), "4E60BC", "22 - Alizarian", False],
    [(237, 226, 237), "EDE2ED", "23 - Apple Blossom", False],
    [(224, 215, 238), "EED7E0", "24 - White Lavender", False],
    [(218, 210, 233), "E9D2DA", "25 - Lavender - Ultra Light", False],
    [(207, 200, 222), "DEC8CF", "26 - Lavender - Pale", False],
    [(233, 236, 252), "FCECE9", "27 - White Violet", False],
    [(125, 78, 146), "924E7D", "28 - Eggplant - Medium Light", False],
    [(103, 64, 118), "764067", "29 - Eggplant", False],
    [(109, 84, 211), "D3546D", "30 - Blueberry - Medium Light", False],
    [(88, 52, 163), "A33458", "31 - Blueberry", False],
    [(77, 46, 138), "8A2E4D", "32 - Blueberry - Dark", False],
    [(217, 83, 159), "9F53D9", "33 - Fuchsia", False],
    [(174, 66, 128), "8042AE", "34 - Fuchsia - Dark", False],
    [(115, 43, 85), "552B73", "35 - Fuchsia - Very Dark", False],
    [(207, 0, 83), "5300CF", "150 - Red - Bright", False],
    [(255, 203, 215), "D7CBFF", "151 - Pink", False],
    [(225, 161, 161), "A1A1E1", "152 - Tawny - Dark", False],
    [(234, 197, 235), "EBC5EA", "153 - Lilac", False],
    [(75, 35, 58), "3A234B", "154 - Red - Very Dark", False],
    [(151, 116, 182), "B67497", "155 - Forget-me-not Blue", False],
    [(133, 119, 180), "B47785", "156 - Blue - Medium", False],
    [(181, 184, 234), "EAB8B5", "157 - Blue - Light", False],
    [(57, 48, 104), "683039", "158 - Blue - Dark", False],
    [(188, 181, 222), "DEB5BC", "159 - Petrol Blue - Light", False],
    [(129, 120, 169), "A97881", "160 - Petrol Blue - Medium", False],
    [(96, 86, 139), "8B5660", "161 - Petrol Blue - Dark", False],
    [(202, 231, 240), "F0E7CA", "162 - Baby Blue - Light", False],
    [(85, 122, 96), "607A55", "163 - Green", False],
    [(186, 228, 182), "B6E4BA", "164 - Green - Light", True],
    [(225, 244, 119), "77F4E1", "165 - Green - Bright", False],
    [(173, 194, 56), "38C2AD", "166 - Lime Green", False],
    [(133, 93, 49), "315D85", "167 - Khaki Brown", False],
    [(177, 174, 183), "B7AEB1", "168 - Silver Gray", False],
    [(130, 125, 125), "7D7D82", "169 - Pewter Gray", False],
    [(148, 66, 167), "A74294", "208 - Lavender - Very Dark", False],
    [(186, 114, 198), "C672BA", "209 - Lavender - Dark", False],
    [(212, 159, 225), "E19FD4", "210 - Lavender - Medium", False],
    [(229, 189, 237), "EDBDE5", "211 - Lavender - Light", False],
    [(121, 38, 49), "312679", "221 - Shell Pink - Very Dark", False],
    [(187, 104, 100), "6468BB", "223 - Shell Pink - Light", False],
    [(226, 165, 152), "98A5E2", "224 - Shell Pink - Very Light", False],
    [(248, 217, 205), "CDD9F8", "225 - Shell Pink - Ultra Very Light", True],
    [(108, 49, 22), "16316C", "300 - Mahogany - Very Dark", True],
    [(170, 82, 55), "3752AA", "301 - Mahogany - Medium", False],
    [(161, 12, 57), "390CA1", "304 - Red - Medium", True],
    [(253, 233, 73), "49E9FD", "307 - Lemon", False],
    [(186, 32, 68), "4420BA", "309 - Rose - Dark", False],
    [(0, 0, 0), "000000", "310 - Black", False],
    [(0, 42, 100), "642A00", "311 - Blue - Medium", False],
    [(31, 50, 121), "79321F", "312 - Baby Blue - Very Dark", False],
    [(125, 66, 70), "46427D", "315 - Antique Mauve - Medium Dark", False],
    [(188, 117, 127), "7F75BC", "316 - Antique Mauve - Medium", False],
    [(109, 100, 105), "69646D", "317 - Pewter Gray", False],
    [(153, 155, 157), "9D9B99", "318 - Steel Gray - Light", False],
    [(58, 85, 59), "3B553A", "319 - Pistachio Green - Very Dark", False],
    [(96, 140, 89), "598C60", "320 - Pistachio Green - Medium", False],
    [(189, 17, 54), "3611BD", "321 - Red", False],
    [(58, 96, 157), "9D603A", "322 - Baby Blue", False],
    [(172, 28, 55), "371CAC", "326 - Rose - Very Dark", False],
    [(94, 15, 119), "770F5E", "327 - Violet", True],
    [(110, 46, 155), "9B2E6E", "333 - Blue Violet - Very Dark", False],
    [(96, 133, 184), "B88560", "334 - Baby Blue - Medium", False],
    [(214, 61, 87), "573DD6", "335 - Rose", False],
    [(12, 39, 94), "5E270C", "336 - Blue", False],
    [(153, 109, 195), "C36D99", "340 - Blue Violet - Medium", False],
    [(163, 154, 215), "D79AA3", "341 - Blue Violet - Light", False],
    [(171, 27, 51), "331BAB", "347 - Salmon - Very Dark", False],
    [(198, 44, 56), "382CC6", "349 - Coral - Dark", False],
    [(222, 63, 64), "403FDE", "350 - Coral - Medium", True],
    [(237, 98, 91), "5B62ED", "351 - Coral", False],
    [(247, 131, 114), "7283F7", "352 - Coral - Light", False],
    [(253, 180, 161), "A1B4FD", "353 - Peach", False],
    [(151, 56, 43), "2B3897", "355 - Terra Cotta - Dark", False],
    [(190, 92, 75), "4B5CBE", "356 - Terra Cotta - Medium", False],
    [(68, 107, 69), "456B44", "367 - Pistachio Green - Dark", False],
    [(127, 198, 109), "6DC67F", "368 - Pistachio Green - Light", False],
    [(205, 239, 166), "A6EFCD", "369 - Pistachio Green - Very Light", False],
    [(145, 114, 69), "457291", "370 - Mustard - Medium", False],
    [(159, 131, 82), "52839F", "371 - Mustard", False],
    [(173, 149, 100), "6495AD", "372 - Mustard - Light", False],
    [(129, 55, 24), "183781", "400 - Mahogany - Dark", False],
    [(239, 158, 116), "749EEF", "402 - Mahogany - Very Light", False],
    [(183, 113, 89), "5971B7", "407 - Desert Sand - Dark", False],
    [(74, 71, 73), "49474A", "413 - Pewter Gray - Dark", False],
    [(118, 110, 114), "726E76", "414 - Steel Gray - Dark", False],
    [(184, 185, 189), "BDB9B8", "415 - Pearl Gray", False],
    [(133, 90, 48), "305A85", "420 - Hazelnut Brown - Dark", False],
    [(201, 154, 103), "679AC9", "422 - Hazelnut Brown - Light", False],
    [(115, 66, 30), "1E4273", "433 - Brown - Medium", False],
    [(143, 83, 50), "32538F", "434 - Brown - Light", False],
    [(169, 101, 56), "3865A9", "435 - Brown - Very Light", False],
    [(199, 133, 89), "5985C7", "436 - Tan", False],
    [(218, 162, 111), "6FA2DA", "437 - Tan - Light", False],
    [(245, 188, 19), "13BCF5", "444 - Lemon - Dark", False],
    [(252, 249, 153), "99F9FC", "445 - Lemon - Light", False],
    [(136, 119, 115), "737788", "451 - Shell Gray - Dark", False],
    [(173, 153, 148), "9499AD", "452 - Shell Gray - Medium", False],
    [(204, 184, 170), "AAB8CC", "453 - Shell Gray - Light", False],
    [(91, 101, 51), "33655B", "469 - Avocado Green", False],
    [(114, 129, 62), "3E8172", "470 - Avocado Green - Light", False],
    [(158, 179, 87), "57B39E", "471 - Avocado Green - Very Light", True],
    [(209, 222, 117), "75DED1", "472 - Avocado Green - Ultra Light", False],
    [(151, 11, 44), "2C0B97", "498 - Red - Dark", False],
    [(29, 54, 42), "2A361D", "500 - Blue Green - Very Dark", False],
    [(47, 84, 70), "46542F", "501 - Blue Green - Dark", True],
    [(87, 130, 110), "6E8257", "502 - Blue Green", False],
    [(137, 184, 159), "9FB889", "503 - Blue Green - Medium", False],
    [(51, 131, 98), "628333", "505 - Grass Green - Dark", False],
    [(33, 98, 133), "856221", "517 - Wedgewood - Dark", False],
    [(80, 129, 156), "9C8150", "518 - Wedgewood - Light", False],
    [(148, 183, 203), "CBB794", "519 - Sky Blue", False],
    [(56, 69, 38), "264538", "520 - Fern Green - Dark", True],
    [(128, 139, 110), "6E8B80", "522 - Fern Green", False],
    [(149, 159, 122), "7A9F95", "523 - Fern Green - Light", False],
    [(174, 167, 142), "8EA7AE", "524 - Fern Green - Very Light", False],
    [(75, 75, 73), "494B4B", "535 - Ash Gray - Very Light", False],
    [(234, 208, 181), "B5D0EA", "543 - Beige Brown - Ultra Very Light", False],
    [(88, 14, 92), "5C0E58", "550 - Violet - Very Dark", False],
    [(144, 47, 153), "992F90", "552 - Violet - Medium", False],
    [(164, 73, 172), "AC49A4", "553 - Violet", False],
    [(220, 156, 222), "DE9CDC", "554 - Violet - Light", False],
    [(40, 94, 72), "485E28", "561 - Jade - Very Dark", False],
    [(59, 140, 90), "5A8C3B", "562 - Jade - Medium", False],
    [(110, 211, 154), "9AD36E", "563 - Jade - Light", False],
    [(149, 228, 175), "AFE495", "564 - Jade - Very Light", False],
    [(53, 95, 11), "0B5F35", "580 - Moss Green - Dark", False],
    [(131, 138, 41), "298A83", "581 - Moss Green", False],
    [(82, 173, 171), "ABAD52", "597 - Turquoise", False],
    [(151, 216, 211), "D3D897", "598 - Turquoise - Light", False],
    [(191, 28, 72), "481CBF", "600 - Cranberry - Very Dark", False],
    [(198, 42, 83), "532AC6", "601 - Cranberry - Dark", False],
    [(214, 63, 104), "683FD6", "602 - Cranberry - Medium", False],
    [(251, 75, 124), "7C4BFB", "603 - Cranberry - Light Medium", False],
    [(247, 147, 178), "B293F7", "604 - Cranberry - Light", False],
    [(251, 172, 196), "C4ACFB", "605 - Cranberry - Very Light", False],
    [(247, 15, 0), "000FF7", "606 - Orange-red - Bright", False],
    [(253, 72, 12), "0C48FD", "608 - Orange - Bright", False],
    [(107, 80, 57), "39506B", "610 - Drab Brown - Dark", False],
    [(124, 95, 70), "465F7C", "611 - Drab Brown", False],
    [(166, 136, 94), "5E88A6", "612 - Drab Brown - Light", False],
    [(185, 159, 114), "729FB9", "613 - Drab Brown - Very Light", False],
    [(127, 66, 50), "32427F", "632 - Desert Sand - Ultra Very Dark", False],
    [(129, 120, 104), "687881", "640 - Beige Gray - Very Dark", False],
    [(149, 141, 121), "798D95", "642 - Beige Gray - Dark", False],
    [(196, 190, 166), "A6BEC4", "644 - Beige Gray - Medium", False],
    [(93, 93, 84), "545D5D", "645 - Beaver Gray - Very Dark", False],
    [(107, 104, 96), "60686B", "646 - Beaver Gray - Dark", False],
    [(144, 142, 133), "858E90", "647 - Beaver Gray - Medium", False],
    [(167, 166, 159), "9FA6A7", "648 - Beaver Gray - Light", False],
    [(206, 27, 51), "331BCE", "666 - Red - Bright", False],
    [(236, 191, 125), "7DBFEC", "676 - Old Gold - Light", False],
    [(242, 220, 159), "9FDCF2", "677 - Old Gold - Very Light", False],
    [(176, 123, 70), "467BB0", "680 - Old Gold - Dark", False],
    [(7, 91, 38), "265B07", "699 - Green", False],
    [(7, 108, 52), "346C07", "700 - Green - Bright", False],
    [(33, 124, 54), "367C21", "701 - Green - Light", False],
    [(55, 145, 48), "309137", "702 - Kelly Green", False],
    [(99, 179, 48), "30B363", "703 - Chartreuse", False],
    [(136, 197, 58), "3AC588", "704 - Chartreuse - Bright", False],
    [(246, 239, 218), "DAEFF6", "712 - Cream", False],
    [(203, 32, 137), "8920CB", "718 - Plum", False],
    [(200, 58, 36), "243AC8", "720 - Orange Spice - Dark", True],
    [(244, 100, 64), "4064F4", "721 - Orange Spice - Medium", False],
    [(249, 135, 86), "5687F9", "722 - Orange Spice - Light", False],
    [(249, 208, 57), "39D0F9", "724 - Winnie The Pooh Gold", False],
    [(249, 193, 91), "5BC1F9", "725 - Topaz", False],
    [(253, 219, 99), "63DBFD", "726 - Topaz - Light", False],
    [(253, 233, 139), "8BE9FD", "727 - Topaz - Very Light", False],
    [(242, 174, 63), "3FAEF2", "728 - Golden Yellow", False],
    [(206, 150, 87), "5796CE", "729 - Old Gold - Medium", False],
    [(99, 82, 11), "0B5263", "730 - Olive Green - Very Dark", False],
    [(114, 92, 12), "0C5C72", "732 - Olive Green", False],
    [(167, 138, 68), "448AA7", "733 - Olive Green - Medium", False],
    [(187, 156, 84), "549CBB", "734 - Olive Green - Light", False],
    [(226, 183, 131), "83B7E2", "738 - Tan - Very Light", False],
    [(242, 222, 185), "B9DEF2", "739 - Tan - Ultra Very Light", False],
    [(253, 111, 26), "1A6FFD", "740 - Tangerine", False],
    [(252, 139, 16), "108BFC", "741 - Tangerine - Medium", True],
    [(253, 174, 60), "3CAEFD", "742 - Tangerine - Light", False],
    [(253, 215, 105), "69D7FD", "743 - Yellow - Medium", False],
    [(254, 232, 141), "8DE8FE", "744 - Yellow - Pale", False],
    [(254, 235, 165), "A5EBFE", "745 - Yellow - Light Pale", False],
    [(250, 242, 213), "D5F2FA", "746 - Off White", False],
    [(206, 233, 234), "EAE9CE", "747 - Sky Blue - Very Light", False],
    [(247, 201, 176), "B0C9F7", "754 - Peach - Light", False],
    [(233, 159, 131), "839FE9", "758 - Terra Cotta - Very Light", False],
    [(236, 136, 128), "8088EC", "760 - Salmon", False],
    [(248, 180, 173), "ADB4F8", "761 - Salmon - Light", False],
    [(209, 208, 210), "D2D0D1", "762 - Pearl Gray - Very Light", False],
    [(215, 239, 167), "A7EFD7", "772 - Yellow Green - Very Light", True],
    [(212, 227, 239), "EFE3D4", "775 - Baby Blue - Very Light", False],
    [(155, 0, 66), "42009B", "777 - Red - Deep", False],
    [(220, 166, 164), "A4A6DC", "778 - Antique Mauve - Very Light", False],
    [(83, 51, 45), "2D3353", "779 - Brown", False],
    [(148, 80, 38), "265094", "780 - Topaz - Ultra Very Dark", False],
    [(178, 105, 35), "2369B2", "782 - Topaz - Dark", False],
    [(208, 136, 61), "3D88D0", "783 - Topaz - Medium", True],
    [(45, 32, 104), "68202D", "791 - Cornflower Blue - Very Dark", False],
    [(69, 75, 139), "8B4B45", "792 - Cornflower Blue - Dark", False],
    [(124, 130, 181), "B5827C", "793 - Cornflower Blue - Medium", False],
    [(160, 178, 215), "D7B2A0", "794 - Cornflower Blue - Light", False],
    [(39, 34, 118), "762227", "796 - Royal Blue - Dark", False],
    [(43, 50, 136), "88322B", "797 - Royal Blue", False],
    [(78, 92, 167), "A75C4E", "798 - Delft Blue - Dark", False],
    [(107, 127, 192), "C07F6B", "799 - Delft Blue - Medium", False],
    [(181, 199, 233), "E9C7B5", "800 - Delft Blue - Pale", False],
    [(96, 57, 29), "1D3960", "801 - Coffee Brown - Dark", False],
    [(32, 39, 84), "542720", "803 - Blue - Deep", False],
    [(85, 139, 158), "9E8B55", "807 - Peacock Blue", False],
    [(145, 159, 213), "D59F91", "809 - Delft Blue", False],
    [(127, 160, 198), "C6A07F", "813 - Blue - Light", False],
    [(113, 16, 51), "331071", "814 - Garnet - Dark", False],
    [(128, 11, 52), "340B80", "815 - Garnet - Medium", False],
    [(146, 18, 56), "381292", "816 - Garnet", False],
    [(187, 22, 48), "3016BB", "817 - Coral Red - Very Dark", False],
    [(254, 222, 221), "DDDEFE", "818 - Baby Pink", False],
    [(252, 235, 222), "DEEBFC", "819 - Baby Pink - Light", False],
    [(21, 18, 100), "641215", "820 - Royal Blue - Very Dark", False],
    [(232, 223, 199), "C7DFE8", "822 - Beige Gray - Light", False],
    [(0, 11, 68), "440B00", "823 - Blue - Dark", False],
    [(40, 71, 121), "794728", "824 - Blue - Very Dark", True],
    [(52, 88, 143), "8F5834", "825 - Blue - Dark", False],
    [(80, 117, 167), "A77550", "826 - Blue - Medium", False],
    [(164, 193, 222), "DEC1A4", "827 - Blue - Very Light", False],
    [(195, 215, 230), "E6D7C3", "828 - Blue - Ultra Very Light", False],
    [(100, 72, 12), "0C4864", "829 - Golden Olive - Very Dark", False],
    [(110, 80, 29), "1D506E", "830 - Golden Olive - Dark", False],
    [(124, 95, 32), "205F7C", "831 - Golden Olive - Medium", False],
    [(156, 114, 48), "30729C", "832 - Golden Olive", False],
    [(185, 153, 86), "5699B9", "833 - Golden Olive - Light", False],
    [(210, 180, 104), "68B4D2", "834 - Golden Olive - Very Light", False],
    [(74, 48, 33), "21304A", "838 - Beige Brown - Very Dark", False],
    [(90, 60, 45), "2D3C5A", "839 - Beige Brown - Dark", False],
    [(122, 89, 57), "39597A", "840 - Beige Brown - Medium", False],
    [(163, 125, 100), "647DA3", "841 - Beige Brown - Light", False],
    [(203, 176, 148), "94B0CB", "842 - Beige Brown - Very Light", False],
    [(73, 72, 66), "424849", "844 - Beaver Gray - Ultra Dark", False],
    [(120, 76, 40), "284C78", "869 - Hazelnut Brown - Very Dark", False],
    [(50, 66, 51), "334232", "890 - Pistachio Green - Ultra Dark", False],
    [(238, 50, 70), "4632EE", "891 - Carnation - Dark", False],
    [(244, 71, 83), "5347F4", "892 - Carnation - Medium", False],
    [(246, 104, 121), "7968F6", "893 - Carnation - Light", False],
    [(253, 149, 163), "A395FD", "894 - Carnation - Very Light", False],
    [(52, 75, 46), "2E4B34", "895 - Hunter Green - Very Dark", False],
    [(83, 47, 27), "1B2F53", "898 - Coffee Brown - Very Dark", False],
    [(234, 107, 120), "786BEA", "899 - Rose - Medium", False],
    [(198, 49, 23), "1731C6", "900 - Burnt Orange - Dark", False],
    [(101, 19, 41), "291365", "902 - Garnet - Very Dark", False],
    [(56, 99, 36), "246338", "904 - Parrot Green - Very Dark", False],
    [(70, 121, 36), "247946", "905 - Parrot Green - Dark", False],
    [(108, 158, 41), "299E6C", "906 - Parrot Green - Medium", False],
    [(157, 199, 45), "2DC79D", "907 - Parrot Green - Light", False],
    [(16, 107, 67), "436B10", "909 - Emerald Green - Very Dark", False],
    [(16, 129, 78), "4E8110", "910 - Emerald Green - Dark", False],
    [(16, 146, 86), "569210", "911 - Emerald Green - Medium", False],
    [(54, 178, 107), "6BB236", "912 - Emerald Green - Light", False],
    [(85, 202, 125), "7DCA55", "913 - Nile Green - Medium", False],
    [(149, 8, 90), "5A0895", "915 - Plum - Dark", True],
    [(172, 16, 113), "7110AC", "917 - Plum - Medium", False],
    [(136, 54, 48), "303688", "918 - Red Copper - Dark", False],
    [(155, 55, 27), "1B379B", "919 - Red Copper", False],
    [(171, 72, 54), "3648AB", "920 - Copper - Medium", False],
    [(192, 87, 61), "3D57C0", "921 - Copper", False],
    [(221, 110, 76), "4C6EDD", "922 - Copper - Light", False],
    [(56, 74, 74), "4A4A38", "924 - Gray Green - Very Dark", False],
    [(97, 118, 116), "747661", "926 - Gray Green - Medium", False],
    [(159, 168, 165), "A5A89F", "927 - Gray Green - Light", False],
    [(192, 198, 192), "C0C6C0", "928 - Gray Green - Very Light", False],
    [(73, 92, 107), "6B5C49", "930 - Antique Blue - Dark", False],
    [(102, 118, 132), "847666", "931 - Antique Blue - Medium", False],
    [(147, 160, 175), "AFA093", "932 - Antique Blue - Light", False],
    [(50, 51, 36), "243332", "934 - Avocado Green - BLACK", False],
    [(56, 58, 42), "2A3A38", "935 - Avocado Green - Dark", False],
    [(63, 66, 39), "27423F", "936 - Avocado Green - Very Dark", False],
    [(67, 79, 44), "2C4F43", "937 - Avocado Green - Medium", False],
    [(69, 39, 26), "1A2745", "938 - Coffee Brown - Ultra Dark", False],
    [(9, 9, 47), "2F0909", "939 - Blue - Very Dark", False],
    [(0, 154, 119), "779A00", "943 - Aquamarine - Medium", False],
    [(246, 193, 154), "9AC1F6", "945 - Tawny", False],
    [(237, 65, 21), "1541ED", "946 - Burnt Orange - Medium", False],
    [(252, 79, 22), "164FFC", "947 - Burnt Orange", False],
    [(253, 230, 211), "D3E6FD", "948 - Peach - Very Light", False],
    [(229, 172, 141), "8DACE5", "950 - Desert Sand - Light", False],
    [(250, 221, 182), "B6DDFA", "951 - Tawny - Light", False],
    [(111, 218, 138), "8ADA6F", "954 - Nile Green", False],
    [(168, 235, 173), "ADEBA8", "955 - Nile Green - Light", False],
    [(247, 86, 109), "6D56F7", "956 - Geranium", False],
    [(253, 153, 175), "AF99FD", "957 - Geranium - Pale", True],
    [(13, 178, 148), "94B20D", "958 - Seagreen - Dark", False],
    [(114, 208, 183), "B7D072", "959 - Seagreen - Medium", False],
    [(222, 88, 108), "6C58DE", "961 - Dusty Rose - Dark", False],
    [(235, 113, 131), "8371EB", "962 - Dusty Rose - Medium", False],
    [(253, 204, 209), "D1CCFD", "963 - Dusty Rose - Ultra Very Light", False],
    [(165, 228, 212), "D4E4A5", "964 - Seagreen - Light", True],
    [(148, 210, 138), "8AD294", "966 - Baby Green - Medium", False],
    [(255, 194, 172), "ACC2FF", "967 - Peach - Light", False],
    [(251, 103, 33), "2167FB", "970 - Pumpkin - Light", False],
    [(251, 159, 17), "119FFB", "972 - Canary - Deep", False],
    [(252, 205, 45), "2DCDFC", "973 - Canary - Bright", False],
    [(249, 199, 57), "39C7F9", "974 - Winnie The Pooh Gold 2", False],
    [(129, 60, 17), "113C81", "975 - Golden Brown - Dark", False],
    [(207, 117, 50), "3275CF", "976 - Golden Brown - Medium", False],
    [(236, 143, 67), "438FEC", "977 - Golden Brown - Light", False],
    [(46, 82, 48), "30522E", "986 - Forest Green - Very Dark", False],
    [(67, 104, 56), "386843", "987 - Forest Green - Dark", False],
    [(102, 146, 74), "4A9266", "988 - Forest Green - Medium", True],
    [(113, 167, 78), "4EA771", "989 - Forest Green", False],
    [(19, 95, 85), "555F13", "991 - Aquamarine - Dark", False],
    [(66, 181, 158), "9EB542", "992 - Aquamarine - Light", False],
    [(98, 216, 182), "B6D862", "993 - Aquamarine - Very Light", False],
    [(0, 97, 176), "B06100", "995 - Electric Blue - Dark", False],
    [(73, 168, 235), "EBA849", "996 - Electric Blue - Medium", False],
    [(101, 89, 53), "355965", "3011 - Khaki Green - Dark", False],
    [(139, 123, 78), "4E7B8B", "3012 - Khaki Green - Medium", False],
    [(175, 169, 123), "7BA9AF", "3013 - Khaki Green - Light", False],
    [(80, 64, 59), "3B4050", "3021 - Brown Gray - Very Dark", False],
    [(132, 130, 116), "748284", "3022 - Brown Gray - Medium", False],
    [(162, 155, 134), "869BA2", "3023 - Brown Gray - Light", False],
    [(190, 184, 172), "ACB8BE", "3024 - Brown Gray - Very Light", False],
    [(66, 48, 20), "143042", "3031 - Mocha Brown - Very Dark", False],
    [(157, 136, 104), "68889D", "3032 - Mocha Brown - Medium", False],
    [(219, 199, 173), "ADC7DB", "3033 - Mocha Brown - Very Light", False],
    [(134, 106, 118), "766A86", "3041 - Antique Violet - Medium", False],
    [(175, 152, 160), "A098AF", "3042 - Antique Violet - Light", False],
    [(175, 129, 82), "5281AF", "3045 - Yellow Beige - Dark", False],
    [(206, 176, 116), "74B0CE", "3046 - Yellow Beige - Medium", False],
    [(234, 216, 171), "ABD8EA", "3047 - Yellow Beige - Light", False],
    [(76, 76, 30), "1E4C4C", "3051 - Green Gray - Dark", False],
    [(120, 126, 92), "5C7E78", "3052 - Green Gray - Medium", False],
    [(153, 157, 117), "759D99", "3053 - Green Gray", False],
    [(186, 112, 86), "5670BA", "3064 - Desert Sand", False],
    [(210, 210, 202), "CAD2D2", "3072 - Beaver Gray - Very Light", False],
    [(252, 246, 182), "B6F6FC", "3078 - Golden Yellow - Very Light", False],
    [(173, 205, 231), "E7CDAD", "3325 - Baby Blue - Light", False],
    [(249, 151, 156), "9C97F9", "3326 - Rose - Light", False],
    [(190, 68, 74), "4A44BE", "3328 - Salmon - Dark", False],
    [(253, 107, 79), "4F6BFD", "3340 - Apricot - Medium", False],
    [(253, 142, 120), "788EFD", "3341 - Apricot", False],
    [(64, 85, 46), "2E5540", "3345 - Hunter Green - Dark", False],
    [(86, 116, 59), "3B7456", "3346 - Hunter Green", False],
    [(109, 150, 70), "46966D", "3347 - Yellow Green - Medium", False],
    [(190, 223, 116), "74DFBE", "3348 - Yellow Green - Light", False],
    [(170, 57, 73), "4939AA", "3350 - Dusty Rose - Ultra Dark", False],
    [(239, 165, 172), "ACA5EF", "3354 - Dusty Rose - Light", True],
    [(73, 82, 60), "3C5249", "3362 - Pine Green - Dark", False],
    [(97, 116, 81), "517461", "3363 - Pine Green - Medium", False],
    [(142, 155, 109), "6D9B8E", "3364 - Pine Green", False],
    [(54, 34, 14), "0E2236", "3371 - Black Brown", False],
    [(217, 76, 157), "9D4CD9", "3607 - Plum - Light", False],
    [(236, 129, 190), "BE81EC", "3608 - Plum - Very Light", False],
    [(246, 176, 223), "DFB0F6", "3609 - Plum - Ultra Light", False],
    [(121, 38, 59), "3B2679", "3685 - Mauve - Very Dark", False],
    [(181, 69, 93), "5D45B5", "3687 - Mauve", True],
    [(220, 124, 134), "867CDC", "3688 - Mauve - Medium", False],
    [(248, 187, 200), "C8BBF8", "3689 - Mauve - Light", False],
    [(242, 73, 79), "4F49F2", "3705 - Melon - Dark", False],
    [(253, 110, 112), "706EFD", "3706 - Melon - Medium", False],
    [(253, 160, 174), "AEA0FD", "3708 - Melon - Light", False],
    [(217, 93, 93), "5D5DD9", "3712 - Salmon - Medium", False],
    [(253, 213, 208), "D0D5FD", "3713 - Salmon - Very Light", False],
    [(252, 175, 185), "B9AFFC", "3716 - Dusty Rose - Very Light", False],
    [(147, 59, 61), "3D3B93", "3721 - Shell Pink - Dark", False],
    [(160, 75, 76), "4C4BA0", "3722 - Shell Pink - Medium", False],
    [(149, 86, 92), "5C5695", "3726 - Antique Mauve - Dark", False],
    [(218, 158, 166), "A69EDA", "3727 - Antique Mauve - Light", False],
    [(195, 76, 92), "5C4CC3", "3731 - Dusty Rose - Very Dark", False],
    [(234, 126, 134), "867EEA", "3733 - Dusty Rose", True],
    [(113, 83, 93), "5D5371", "3740 - Antique Violet - Dark", False],
    [(207, 194, 201), "C9C2CF", "3743 - Antique Violet - Very Light", False],
    [(132, 74, 181), "B54A84", "3746 - Blue Violet - Dark", False],
    [(208, 197, 236), "ECC5D0", "3747 - Blue Violet - Very Light", False],
    [(29, 69, 82), "52451D", "3750 - Antique Blue - Very Dark", False],
    [(186, 201, 204), "CCC9BA", "3752 - Antique Blue - Very Light", False],
    [(217, 230, 236), "ECE6D9", "3753 - Antique Blue - Ultra Very Light", False],
    [(129, 165, 216), "D8A581", "3755 - Baby Blue", False],
    [(233, 244, 250), "FAF4E9", "3756 - Baby Blue - Light", False],
    [(70, 114, 147), "937246", "3760 - Wedgewood - Medium", False],
    [(177, 208, 223), "DFD0B1", "3761 - Sky Blue - Light", False],
    [(23, 94, 120), "785E17", "3765 - Peacock Blue - Very Dark", False],
    [(75, 138, 161), "A18A4B", "3766 - Peacock Blue - Light", False],
    [(76, 96, 95), "5F604C", "3768 - Gray Green - Dark", False],
    [(254, 241, 216), "D8F1FE", "3770 - Tawny - Very Light", False],
    [(232, 172, 155), "9BACE8", "3771 - Peach - Dark", False],
    [(153, 87, 68), "445799", "3772 - Desert Sand - Very Dark", False],
    [(243, 207, 180), "B4CFF3", "3774 - Desert Sand - Very Light", False],
    [(201, 100, 68), "4464C9", "3776 - Mahogany - Light", False],
    [(146, 47, 37), "252F92", "3777 - Terra Cotta - Very Dark", False],
    [(210, 112, 92), "5C70D2", "3778 - Terra Cotta - Light", False],
    [(242, 171, 149), "95ABF2", "3779 - Terra Cotta - Ultra Very Light", False],
    [(89, 63, 43), "2B3F59", "3781 - Mocha Brown - Dark", False],
    [(182, 157, 128), "809DB6", "3782 - Mocha Brown - Light", False],
    [(98, 82, 76), "4C5262", "3787 - Brown Gray - Dark", False],
    [(109, 90, 75), "4B5A6D", "3790 - Beige Gray - Ultra Dark", False],
    [(57, 57, 61), "3D3939", "3799 - Pewter Gray - Very Dark", False],
    [(228, 53, 61), "3D35E4", "3801 - Melon - Very Dark", False],
    [(103, 42, 51), "332A67", "3802 - Antique Mauve - Very Dark", False],
    [(135, 42, 67), "432A87", "3803 - Mauve - Dark", False],
    [(206, 43, 99), "632BCE", "3804 - Cyclamen Pink - Dark", False],
    [(223, 60, 115), "733CDF", "3805 - Cyclamen Pink", False],
    [(241, 90, 145), "915AF1", "3806 - Cyclamen Pink - Light", False],
    [(75, 89, 158), "9E594B", "3807 - Cornflower Blue", False],
    [(3, 83, 92), "5C5303", "3808 - Turquoise - Ultra Very Dark", False],
    [(19, 106, 117), "756A13", "3809 - Turquoise - Very Dark", False],
    [(45, 141, 152), "988D2D", "3810 - Turquoise - Dark", False],
    [(168, 226, 229), "E5E2A8", "3811 - Turquoise - Very Light", False],
    [(7, 161, 132), "84A107", "3812 - Seagreen - Very Dark", False],
    [(134, 195, 171), "ABC386", "3813 - Blue Green - Light", False],
    [(11, 134, 115), "73860B", "3814 - Aquamarine", False],
    [(67, 114, 89), "597243", "3815 - Celadon Green - Dark", False],
    [(96, 147, 122), "7A9360", "3816 - Celadon Green", False],
    [(129, 198, 164), "A4C681", "3817 - Celadon Green - Light", False],
    [(0, 93, 46), "2E5D00", "3818 - Emerald Green - Ultra Very Dark", False],
    [(204, 201, 89), "59C9CC", "3819 - Moss Green - Light", False],
    [(219, 165, 62), "3EA5DB", "3820 - Straw - Dark", False],
    [(235, 187, 82), "52BBEB", "3821 - Straw", False],
    [(247, 209, 105), "69D1F7", "3822 - Straw - Light", False],
    [(254, 245, 205), "CDF5FE", "3823 - Yellow - Ultra Pale", True],
    [(252, 174, 153), "99AEFC", "3824 - Apricot - Light", False],
    [(254, 163, 112), "70A3FE", "3825 - Pumpkin - Pale", True],
    [(177, 102, 51), "3366B1", "3826 - Golden Brown", False],
    [(234, 166, 100), "64A6EA", "3827 - Golden Brown - Pale", False],
    [(170, 124, 67), "437CAA", "3828 - Hazelnut Brown", False],
    [(167, 103, 29), "1D67A7", "3829 - Old Gold - Very Dark", False],
    [(169, 65, 56), "3841A9", "3830 - Terra Cotta", False],
    [(193, 43, 82), "522BC1", "3831 - Raspberry - Dark", False],
    [(227, 99, 112), "7063E3", "3832 - Raspberry - Medium", False],
    [(234, 139, 150), "968BEA", "3833 - Raspberry - Light", False],
    [(106, 34, 88), "58226A", "3834 - Grape - Dark", False],
    [(146, 77, 120), "784D92", "3835 - Grape - Medium", False],
    [(197, 151, 185), "B997C5", "3836 - Grape - Light", False],
    [(138, 42, 143), "8F2A8A", "3837 - Lavender - Ultra Dark", False],
    [(96, 107, 173), "AD6B60", "3838 - Lavender Blue - Dark", False],
    [(122, 126, 197), "C57E7A", "3839 - Lavender Blue - Medium", True],
    [(178, 189, 234), "EABDB2", "3840 - Lavender Blue - Light", True],
    [(217, 234, 242), "F2EAD9", "3841 - Baby Blue - Pale", False],
    [(6, 80, 106), "6A5006", "3842 - Wedgewood - Dark", False],
    [(40, 163, 222), "DEA328", "3843 - Electric Blue", False],
    [(31, 127, 160), "A07F1F", "3844 - Bright Turquoise - Dark", False],
    [(43, 173, 209), "D1AD2B", "3845 - Bright Turquoise - Medium", False],
    [(94, 204, 236), "ECCC5E", "3846 - Bright Turquoise - Light", False],
    [(24, 99, 88), "586318", "3847 - Teal Green - Dark", False],
    [(32, 126, 114), "727E20", "3848 - Teal Green - Medium", False],
    [(53, 177, 147), "93B135", "3849 - Teal Green - Light", False],
    [(32, 139, 70), "468B20", "3850 - Bright Green - Dark", False],
    [(97, 187, 132), "84BB61", "3851 - Bright Green - Light", False],
    [(227, 167, 48), "30A7E3", "3852 - Straw - Very Dark", False],
    [(239, 129, 37), "2581EF", "3853 - Autumn Gold - Dark", False],
    [(251, 172, 86), "56ACFB", "3854 - Autumn Gold - Medium", False],
    [(253, 223, 160), "A0DFFD", "3855 - Autumn Gold - Light", False],
    [(253, 190, 142), "8EBEFD", "3856 - Mahogany - Ultra Very Light", False],
    [(106, 47, 38), "262F6A", "3857 - Rosewood - Dark", False],
    [(128, 58, 50), "323A80", "3858 - Rosewood - Medium", False],
    [(186, 122, 108), "6C7ABA", "3859 - Rosewood - Light", False],
    [(137, 99, 98), "626389", "3860 - Cocoa", False],
    [(172, 133, 131), "8385AC", "3861 - Cocoa - Light", False],
    [(110, 73, 42), "2A496E", "3862 - Mocha Beige - Dark", False],
    [(148, 114, 93), "5D7294", "3863 - Mocha Beige - Medium", False],
    [(201, 170, 146), "92AAC9", "3864 - Mocha Beige - Light", False],
    [(255, 253, 249), "F9FDFF", "3865 - Winter White", False],
    [(240, 230, 215), "D7E6F0", "3866 - Mocha Brown - Ultra Very Light", False]
]

# Processed Master List: Filters based on ownership status
if FILTER_UNOWNED_COLORS:
    PROCESSED_MASTER_LIST = [color for color in MASTER_LIST if color[3]]
else:
    PROCESSED_MASTER_LIST = MASTER_LIST.copy()

def find_prominent_colors(image_path, num_colors):
    """
    Find the most prominent colors in the image using k-means clustering.
    """
    image = Image.open(image_path).convert("RGB")
    image_array = np.array(image)
    pixels = image_array.reshape(-1, 3)

    kmeans = KMeans(n_clusters=num_colors, random_state=42)
    kmeans.fit(pixels)
    cluster_centers = kmeans.cluster_centers_

    return cluster_centers

def match_colors_to_master(cluster_centers, master_list):
    """
    Match cluster centers to the closest colors in the master list.
    """
    matched_colors = []
    for center in cluster_centers:
        lab_center = convert_color(sRGBColor(center[0] / 255, center[1] / 255, center[2] / 255), LabColor)
        min_distance = float('inf')
        best_match = None

        for color in master_list:
            master_rgb = color[0]
            master_lab = convert_color(sRGBColor(master_rgb[0] / 255, master_rgb[1] / 255, master_rgb[2] / 255), LabColor)
            distance = delta_e_cie2000(lab_center, master_lab)
            if distance < min_distance:
                min_distance = distance
                best_match = color

        if best_match not in matched_colors:
            matched_colors.append(best_match)

    return matched_colors

# Modify the palette processing logic
if not USE_ALL_COLORS:
    # Find prominent colors in the input image
    cluster_centers = find_prominent_colors(INPUT_PATH, MAX_COLORS)

    # Match the cluster centers to the master list
    PROCESSED_MASTER_LIST = match_colors_to_master(cluster_centers, MASTER_LIST)

# Precompute LAB values for the updated PROCESSED_MASTER_LIST
PRECOMPUTED_LAB = {
    tuple(color[0]): convert_color(sRGBColor(color[0][0] / 255, color[0][1] / 255, color[0][2] / 255), LabColor)
    for color in PROCESSED_MASTER_LIST
}

def closest_color_lab(r, g, b):
    input_color = convert_color(sRGBColor(r / 255, g / 255, b / 255), LabColor)
    min_distance = float('inf')
    best_match = None

    for rgb, lab in PRECOMPUTED_LAB.items():
        distance = delta_e_cie2000(input_color, lab, Kl=1, Kc=1, Kh=2)
        if distance < min_distance:
            min_distance = distance
            best_match = rgb

    return best_match

def create_sheet(input_path, output_path, grid_rows, grid_cols):
    image = Image.open(input_path).convert("RGB")
    image_resized = image.resize((grid_cols, grid_rows))

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cross Stitch"

    used_colors = set()
    for row in range(grid_rows):
        for col in range(grid_cols):
            r, g, b = image_resized.getpixel((col, row))
            closest_rgb = closest_color_lab(r, g, b)

            hex_color = next(color[1] for color in PROCESSED_MASTER_LIST if tuple(color[0]) == closest_rgb)
            name = next(color[2] for color in PROCESSED_MASTER_LIST if tuple(color[0]) == closest_rgb)
            ownership = next(color[3] for color in PROCESSED_MASTER_LIST if tuple(color[0]) == closest_rgb)

            used_colors.add((name, hex_color, ownership))

            cell = sheet.cell(row=row + 1, column=col + 1)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    for col in range(1, grid_cols + 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2.5
    for row in range(1, grid_rows + 1):
        sheet.row_dimensions[row].height = 15

    key_start_col = grid_cols + 2
    sheet.cell(row=1, column=key_start_col).value = "Color Key"
    sheet.cell(row=1, column=key_start_col).font = Font(bold=True)

    key_row = 2
    max_name_length = 0

    for name, hex_color, ownership in sorted(used_colors):
        sheet.cell(row=key_row, column=key_start_col).value = name
        max_name_length = max(max_name_length, len(name))

        key_cell = sheet.cell(row=key_row, column=key_start_col + 1)
        key_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        ownership_value = "Owned" if ownership else "Unowned"
        sheet.cell(row=key_row, column=key_start_col + 2).value = ownership_value
        key_row += 1

    sheet.column_dimensions[openpyxl.utils.get_column_letter(key_start_col)].width = max_name_length + 2

    workbook.save(output_path)
    print(f"Excel file saved: {output_path}")

# Run the program
create_sheet(INPUT_PATH, OUTPUT_PATH, GRID_ROWS, GRID_COLS)

# Calculate the end time and time taken
end = time.time()
length = end - start

# Show the results : this can be altered however you like
print("It took", length, "seconds!")
